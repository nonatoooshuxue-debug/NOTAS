import os 
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import sqlite3 as sq
import psycopg2 as ps
from sqlalchemy import create_engine
from io import BytesIO

if "DATABASE_URL" in st.secrets:
    url_banco = st.secrets["DATABASE_URL"]
else:
    url_banco = "postgresql://postgres:[YOUR-PASSWORD]@db.xobgkyusmybtnpdffoae.supabase.co:5432/postgres"

if url_banco.startswith("postgresql://"):
    url_banco = url_banco.replace("postgresql://", "postgresql+psycopg2://", 1)


bancozin = create_engine(url_banco, pool_pre_ping=True)
st.title("NOTAS")
zezin_frete = st.file_uploader("Arrasta o arquivo devagar ai, irmãozinho", type=["xlsx"])

if zezin_frete is not None:
 abas_base = ["NOTA - CRUZ", "NOTA - ITA", "NOTA - AM"]
 conteudo_arq = BytesIO(zezin_frete.read())
 excel = load_workbook(conteudo_arq)
 for nome_base in abas_base:
    if nome_base in excel.sheetnames:
        aba = excel[nome_base]
        df = pd.read_excel(conteudo_arq, sheet_name=nome_base)
        linhas_negrito = []
        for linha in range(2, aba.max_row + 1):
            celula_cri = aba.cell(row=linha, column=1)
            if celula_cri.font and celula_cri.font.bold:
                linhas_negrito.append(True)
            else:
                linhas_negrito.append(False)
        linhas_negrito = linhas_negrito[:len(df)]
        df_paga = df[linhas_negrito]
        df_npaga = df[[not x for x in linhas_negrito]]
        df_paga["Revenda"] = nome_base
        df_npaga["Revenda"] = nome_base
        if "ITA" in nome_base:
            df_ita = df_paga
            df_ita.to_sql("nota_itapipoca", bancozin, if_exists="replace", index=False)
            df_itanpaga = df_npaga
            df_itanpaga.to_sql("nota_itapipoca_indef", bancozin, if_exists="replace", index=False)
        elif "CRUZ" in nome_base:
            df_cruz = df_paga
            df_cruz.to_sql("nota_cruz", bancozin, if_exists="replace", index=False)
            df_cruznpaga = df_npaga
            df_cruznpaga.to_sql("nota_cruz_indef", bancozin, if_exists="replace", index=False)


aba_paga, aba_naopaga = st.tabs(["NOTAS - PAGAS", "NOTAS - PENDENTES"])
with aba_paga:
    st.subheader("SEGUE AS NOTAS PAGAS/BAIXADAS:")
    tabelas_pagas = ["nota_itapipoca", "nota_cruz"]
    for tabela in tabelas_pagas:
        try:
            df_resultado = pd.read_sql(f"SELECT * FROM {tabela}", bancozin)
        except Exception:
            df_resultado = pd.DataFrame()
        if not df_resultado.empty:
            df_resultado["DATA"] = pd.to_datetime(df_resultado["DATA"])
            data_minima = df_resultado["DATA"].min().date()
            data_maxima = df_resultado["DATA"].max().date()
            periodo_min = st.date_input("De", value=data_minima, key=f"min_{tabela}")
            periodo_max = st.date_input("Até", value=data_maxima, key=f"max_{tabela}")
            df_filtrado = df_resultado[df_resultado["DATA"].dt.date.between(periodo_min, periodo_max)]
            st.dataframe(df_filtrado)
            st.divider()


with aba_naopaga:
    st.subheader("SEGUE AS NOTAS COM STATUS INDEFINIDO:")
    tabelas_pagas = ["nota_itapipoca_indef", "nota_cruz_indef"]
    for tabela in tabelas_pagas:
        try:
            df_resultado = pd.read_sql(f"SELECT * FROM {tabela}", bancozin)
        except Exception:
            df_resultado = pd.DataFrame() 

        if not df_resultado.empty:
            df_resultado["DATA"] = pd.to_datetime(df_resultado["DATA"])
            data_minima = df_resultado["DATA"].min().date()
            data_maxima = df_resultado["DATA"].max().date()
            periodo_min = st.date_input("De", value=data_minima, key=f"min_{tabela}")
            periodo_max = st.date_input("Até", value=data_maxima, key=f"max_{tabela}")
            df_filtrado = df_resultado[df_resultado["DATA"].dt.date.between(periodo_min, periodo_max)]
            st.dataframe(df_filtrado)
            st.divider()


